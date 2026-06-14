"""Build the checked-in Sunzone SKU catalogue from the verified Zoho file."""

from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import load_workbook


SOURCE = Path(
    "/Users/admin/Desktop/Sunzone-Complete-Product-Catalog-With-Photos.xlsx"
)
OUTPUT = Path(__file__).resolve().parents[1] / "sku_catalog.json"


def clean(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def route(item: str, sku: str, source_url: str) -> tuple[str, str]:
    text = f"{item} {sku} {source_url}".casefold()
    sku_text = f"{item} {sku}".casefold()
    if "sport-equipment" in text or "garware net" in text or item == "Netting":
        return "Sports Equipment", "Sports Equipment / Turnkey Facility"
    if any(marker in text for marker in (
        "gym-rubber", "gym-laminate", "astro-turf",
    )):
        if "rubber roll" in text:
            return "Powerful", "Gym Rubber Roll Flooring"
        if "square tile" in text:
            return "Powerful", "Gym Rubber Tiles / Laminate Tiles / Mats"
        return "Powerful", "Gym Astro Turf / Sled Track Turf"
    if any(marker in text for marker in (
        "ecolastic-play", "base-layer-granules", "pu-binders-basf",
    )):
        return "Playful", "EPDM Granules / SBR Granules / PU Binder"
    if any(marker in text for marker in (
        "glory-series", "champion-plus", "play-plus", "power-plus",
        "power-shockpad", "wooden-series",
    )):
        return "Joyful", "PVC / Vinyl Badminton Flooring"
    if "hockey-turf" in text:
        return "Graceful", "FIH Hockey Turf"
    if "cricket" in sku_text:
        return "Graceful", "Artificial Cricket Pitch / Cricket Turf"
    if "tennis" in sku_text:
        return "Graceful", "Multi-Sport Turf / Tennis Turf / Padel Turf"
    if any(marker in text for marker in (
        "multi-sports", "tennis-turf", "turf-track",
    )):
        return "Graceful", "Multi-Sport Turf / Tennis Turf / Padel Turf"
    if any(marker in text for marker in (
        "mapei-adhesive", "turf-in-fills", "turf-shockpad",
        "pu adhesive", "shockpad", "silica sand", "seam tape",
    )):
        return "Graceful", "Turf Accessories"
    if "artificial grass" in text:
        if any(marker in text for marker in (
            "football-turf", "fifa", "rugby", "liga turf",
            "fuerte", "hybrid", "nature", "super turf", "diamond turf",
        )):
            return "Graceful", "Football Turf / Box Cricket Turnkey"
        return "Graceful", "Artificial Grass / Landscape Turf"
    raise ValueError(f"Cannot route SKU: {item} / {sku} / {source_url}")


WEBSITE_SYSTEMS = [
    ("Playful", "EPDM Playground / Wet Pour Flooring", "Ecolastic Play"),
    ("Playful", "EPDM Playground / Wet Pour Flooring", "Ecolastic Pace"),
    ("Playful", "EPDM Playground / Wet Pour Flooring", "Ecolastic Score"),
    ("Playful", "EPDM Playground / Wet Pour Flooring", "EPDM System"),
    ("Graceful", "Football Turf / Box Cricket Turnkey", "Box Cricket Turnkey"),
    ("Graceful", "Artificial Grass / Landscape Turf", "Landscape Artificial Grass"),
    ("Graceful", "Multi-Sport Turf / Tennis Turf / Padel Turf", "Padel Turf"),
    ("Powerful", "Gym PVC Sports Flooring", "Gym PVC Sports Flooring"),
    ("Powerful", "Gym Rubber Tiles / Laminate Tiles / Mats", "Gym Mats"),
    ("Acryplay", "Acrylic Sports Court Systems", "Gemstone Acrylic Court Series"),
    ("Acryplay", "PP Modular Sports Court Tiles", "PP Court Tiles"),
    ("Track & Field", "Synthetic Athletic Track Systems", "Full PUR System"),
    ("Track & Field", "Synthetic Athletic Track Systems", "PU Spray System"),
    ("Track & Field", "Synthetic Athletic Track Systems", "Sandwich System"),
    ("Track & Field", "Synthetic Athletic Track Systems", "Prefabricated Track System"),
    ("Sports Equipment", "Sports Equipment / Turnkey Facility", "Turnkey Sports Equipment"),
    ("Woodplay", "Maple / Hevea Wooden Sports Flooring", "Maple Wooden System"),
    ("Woodplay", "Maple / Hevea Wooden Sports Flooring", "Hevea Wooden System"),
]


def build() -> dict:
    workbook = load_workbook(SOURCE, read_only=True, data_only=True)
    sheet = workbook["Product Photo Catalogue"]
    records = []
    seen = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        item = clean(row[1])
        sku = clean(row[2])
        source_url = clean(row[6])
        vertical, family = route(item, sku, source_url)
        display_name = f"{item} - {sku}"
        key = (vertical.casefold(), display_name.casefold())
        if key in seen:
            continue
        seen.add(key)
        records.append({
            "vertical": vertical,
            "search_family": family,
            "item_name": item,
            "sku": sku,
            "display_name": display_name,
            "source_url": source_url,
            "catalogue_type": "SKU",
        })

    for vertical, family, name in WEBSITE_SYSTEMS:
        key = (vertical.casefold(), name.casefold())
        if key in seen:
            continue
        seen.add(key)
        records.append({
            "vertical": vertical,
            "search_family": family,
            "item_name": name,
            "sku": name,
            "display_name": name,
            "source_url": "https://www.sunzone.in/",
            "catalogue_type": "Website system",
        })

    records.sort(key=lambda item: (
        item["vertical"].casefold(),
        item["search_family"].casefold(),
        item["display_name"].casefold(),
    ))
    return {
        "source": "Sunzone verified Zoho catalogue and sunzone.in systems",
        "sku_count": len(records),
        "skus": records,
    }


def main() -> None:
    data = build()
    OUTPUT.write_text(
        json.dumps(data, ensure_ascii=True, indent=2) + "\n",
        encoding="utf-8",
    )
    print(f"Wrote {data['sku_count']} SKU/system records to {OUTPUT}")


if __name__ == "__main__":
    main()
