"""Build the nationwide rotation from Census of India 2011 town directory."""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook


STATE_NAMES = {
    "01": "Jammu and Kashmir",
    "02": "Himachal Pradesh",
    "03": "Punjab",
    "04": "Chandigarh",
    "05": "Uttarakhand",
    "06": "Haryana",
    "07": "Delhi",
    "08": "Rajasthan",
    "09": "Uttar Pradesh",
    "10": "Bihar",
    "11": "Sikkim",
    "12": "Arunachal Pradesh",
    "13": "Nagaland",
    "14": "Manipur",
    "15": "Mizoram",
    "16": "Tripura",
    "17": "Meghalaya",
    "18": "Assam",
    "19": "West Bengal",
    "20": "Jharkhand",
    "21": "Odisha",
    "22": "Chhattisgarh",
    "23": "Madhya Pradesh",
    "24": "Gujarat",
    "25": "Daman and Diu",
    "26": "Dadra and Nagar Haveli",
    "27": "Maharashtra",
    "28": "Andhra Pradesh and Telangana",
    "29": "Karnataka",
    "30": "Goa",
    "31": "Lakshadweep",
    "32": "Kerala",
    "33": "Tamil Nadu",
    "34": "Puducherry",
    "35": "Andaman and Nicobar Islands",
}

PRIMARY_STATE = "Maharashtra"

MAHARASHTRA_PRIORITY_CITIES = [
    ("Mumbai", "Maharashtra"),
    ("Pune", "Maharashtra"),
    ("Nagpur", "Maharashtra"),
    ("Nashik", "Maharashtra"),
    ("Aurangabad", "Maharashtra"),
]

NATIONAL_PRIORITY_CITIES = [
    ("Delhi", "Delhi"),
    ("Bangalore", "Karnataka"),
    ("Chennai", "Tamil Nadu"),
    ("Hyderabad", "Andhra Pradesh and Telangana"),
    ("Kolkata", "West Bengal"),
    ("Ahmedabad", "Gujarat"),
    ("Jaipur", "Rajasthan"),
    ("Lucknow", "Uttar Pradesh"),
    ("Surat", "Gujarat"),
    ("Indore", "Madhya Pradesh"),
    ("Vadodara", "Gujarat"),
    ("Bhopal", "Madhya Pradesh"),
    ("Visakhapatnam", "Andhra Pradesh and Telangana"),
    ("Patna", "Bihar"),
    ("Coimbatore", "Tamil Nadu"),
    ("Gurgaon", "Haryana"),
    ("Noida", "Uttar Pradesh"),
    ("Chandigarh", "Chandigarh"),
    ("Kochi", "Kerala"),
    ("Guwahati", "Assam"),
    ("Bhubaneswar", "Odisha"),
    ("Thiruvananthapuram", "Kerala"),
    ("Rajkot", "Gujarat"),
]

PRIORITY_CITIES = MAHARASHTRA_PRIORITY_CITIES + NATIONAL_PRIORITY_CITIES


def clean_town_name(value: str) -> str:
    name = re.sub(r"\s*\([^)]*\)", "", value)
    return re.sub(r"\s+", " ", name).strip(" ,")


def build(source: Path, output: Path) -> None:
    workbook = load_workbook(source, read_only=True, data_only=True)
    rows = workbook.active.iter_rows(values_only=True)
    next(rows)
    towns = {}
    for state_code, _, _, location_code, raw_name in rows:
        code = str(location_code or "").zfill(6)
        raw_name = str(raw_name or "").strip()
        is_statutory_town = code.startswith("8")
        is_census_town = "(CT)" in raw_name
        if not raw_name or not (is_statutory_town or is_census_town):
            continue
        name = clean_town_name(raw_name)
        state = STATE_NAMES.get(str(state_code).zfill(2), "India")
        if not name:
            continue
        key = (name.casefold(), state.casefold())
        towns[key] = {"city": name, "state": state}

    for city, state in PRIORITY_CITIES:
        towns[(city.casefold(), state.casefold())] = {
            "city": city,
            "state": state,
        }
    priority_order = {
        (name.casefold(), state.casefold()): index
        for index, (name, state) in enumerate(NATIONAL_PRIORITY_CITIES)
    }
    maharashtra_priority = {
        name.casefold(): index
        for index, (name, _) in enumerate(MAHARASHTRA_PRIORITY_CITIES)
    }
    cities = sorted(
        towns.values(),
        key=lambda item: (
            0 if item["state"] == PRIMARY_STATE else 1,
            (
                maharashtra_priority.get(
                    item["city"].casefold(),
                    len(maharashtra_priority),
                )
                if item["state"] == PRIMARY_STATE
                else priority_order.get(
                    (item["city"].casefold(), item["state"].casefold()),
                    len(priority_order),
                )
            ),
            item["state"].casefold(),
            item["city"].casefold(),
        ),
    )
    output.write_text(
        json.dumps(
            {
                "source": (
                    "Census of India 2011 Town and Village Directory, "
                    "PC11_TV_DIR.xlsx"
                ),
                "source_url": (
                    "https://censusindia.gov.in/nada/index.php/catalog/"
                    "42648/download/46323/PC11_TV_DIR.xlsx"
                ),
                "city_count": len(cities),
                "cities": cities,
            },
            indent=2,
            ensure_ascii=True,
        )
        + "\n",
        encoding="utf-8",
    )
    print(f"Wrote {len(cities)} cities to {output}")


if __name__ == "__main__":
    if len(sys.argv) != 3:
        raise SystemExit(
            "Usage: build_india_city_rotation.py SOURCE.xlsx OUTPUT.json"
        )
    build(Path(sys.argv[1]), Path(sys.argv[2]))
