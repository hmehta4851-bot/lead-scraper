"""Build a deduplicated, vertical-first scraper library from the master file."""

from __future__ import annotations

import json
import re
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook

from config import VERTICALS
from keyword_rules import validate_keyword_library, validate_product_keyword


SOURCE = Path("/Users/admin/Desktop/Sunzone-IndiaMART-Product-Keywords.xlsx")
OUTPUT = Path(__file__).with_name("keyword_library.json")


def normalize(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip()


def configured_products() -> set[tuple[str, str]]:
    return {
        (vertical, product)
        for vertical, cfg in VERTICALS.items()
        for product in cfg["products"]
    }


def build() -> dict[str, dict[str, list[str]]]:
    workbook = load_workbook(SOURCE, read_only=True, data_only=True)
    sheet = workbook["Product Keywords"]
    expected = configured_products()
    routed: dict[str, dict[str, list[str]]] = defaultdict(
        lambda: defaultdict(list)
    )
    globally_seen: set[str] = set()
    discovered: set[tuple[str, str]] = set()
    rejected_by_route: dict[tuple[str, str], int] = defaultdict(int)

    for product, keyword, vertical in sheet.iter_rows(min_row=2, values_only=True):
        product_name = normalize(product)
        phrase = normalize(keyword)
        vertical_name = normalize(vertical)
        route = (vertical_name, product_name)
        if route not in expected or not phrase:
            continue
        discovered.add(route)
        valid, _ = validate_product_keyword(
            vertical_name,
            product_name,
            phrase,
        )
        if not valid:
            rejected_by_route[route] += 1
            continue
        key = phrase.casefold()
        if key in globally_seen:
            continue
        globally_seen.add(key)
        routed[vertical_name][product_name].append(phrase)

    missing = sorted(expected - discovered)
    if missing:
        raise ValueError(f"Master workbook is missing configured products: {missing}")

    library = {
        vertical: dict(products)
        for vertical, products in routed.items()
    }
    errors = validate_keyword_library(library)
    if errors:
        raise ValueError(
            "Generated library failed validation: " + "; ".join(errors[:10])
        )

    rejected_total = sum(rejected_by_route.values())
    print(f"Rejected mismatched or ambiguous keywords: {rejected_total}")
    for (vertical, product), count in sorted(rejected_by_route.items()):
        print(f"  {vertical} / {product}: {count}")
    return library


def main() -> None:
    library = build()
    OUTPUT.write_text(
        json.dumps(library, ensure_ascii=True, indent=2) + "\n",
        encoding="utf-8",
    )
    product_count = sum(len(products) for products in library.values())
    keyword_count = sum(
        len(keywords)
        for products in library.values()
        for keywords in products.values()
    )
    print(f"Verticals: {len(library)}")
    print(f"Products: {product_count}")
    print(f"Globally unique keywords: {keyword_count}")
    for vertical, products in library.items():
        total = sum(map(len, products.values()))
        print(f"{vertical}: {len(products)} products, {total} keywords")


if __name__ == "__main__":
    main()
